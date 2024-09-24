import shutil
import openpyxl as openpyxl
from fastapi import Depends
from sqlalchemy.orm import Session
import environ
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.client_credential import ClientCredential
from pathlib import PurePath
from openpyxl.utils import get_column_letter
import os
from src.dependencies.database_dependency import get_sample_db
from src.domains.sharepoint.sharepoint_interface import ISharepointRepository

class SharepointRepository(ISharepointRepository):
    def __init__(self, db: Session = Depends(get_sample_db)):
        self.db = db

    #region Download Data
    def download_dealer_data(self, folder_name):
        env = environ.Env()
        environ.Env.read_env()

        CLIENT_ID = env('CLIENT_ID')
        CLIENT_SECRET = env('CLIENT_SECRET')
        SHAREPOINT_SITE_NAME = env('SHAREPOINT_SITE_NAME')
        SHAREPOINT_SITE = env('SHAREPOINT_DOMAIN_SITE_URL')
        SHAREPOINT_DOC = env('SHAREPOINT_DOMAIN_SITE_DOCUMENT_PATH')

        credentials = ClientCredential(CLIENT_ID, CLIENT_SECRET)

        try:
            try:
                context = ClientContext(f"{SHAREPOINT_SITE}/{SHAREPOINT_SITE_NAME}").with_credentials(credentials)
            except Exception as e:
                print(e)
                print('Error when authenticating application')
                return 'Error when authenticating application'

            downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads", folder_name)
            
            if not os.path.exists(downloads_folder):
                os.makedirs(downloads_folder)

            root_folder = context.web.get_folder_by_server_relative_url(f"{SHAREPOINT_DOC}/{folder_name}")
            root_folder.expand(["Files", "Folders"]).get().execute_query()

            downloaded_files = []
            for file in root_folder.files:
                print(f"Downloading file: {file.properties['ServerRelativeUrl']} ...")
                download_file_name = os.path.join(downloads_folder, file.name)
                with open(download_file_name, "wb") as local_file:
                    file.download(local_file).execute_query()
                downloaded_files.append(download_file_name)
                print(f"[Ok] file has been downloaded: {download_file_name}")

            return f"{downloads_folder}"

        except Exception as e:
            print(e)
            return 'Error when downloading files'
    #endregion

    #region Excel

    #region Upload
    def upload_excel(self, storage_dir, folder_name):
        env = environ.Env()
        environ.Env.read_env()

        CLIENT_ID = env('CLIENT_ID')
        CLIENT_SECRET = env('CLIENT_SECRET')
        SHAREPOINT_SITE_NAME = env('SHAREPOINT_SITE_NAME')
        SHAREPOINT_SITE = env('SHAREPOINT_DOMAIN_SITE_URL')
        SHAREPOINT_DOC = env('SHAREPOINT_DOMAIN_SITE_DOCUMENT_PATH')

        credentials = ClientCredential(
            CLIENT_ID,
            CLIENT_SECRET
        )

        try:
            context = ClientContext(f"{SHAREPOINT_SITE}/{SHAREPOINT_SITE_NAME}").with_credentials(credentials)
        except:
            return('Error when authenticating application')

        storage_files = os.listdir(storage_dir)

        file_list = []
        for item in storage_files:
            item_path = PurePath(storage_dir, item)
            
            if os.path.isfile(item_path):
                file_list.append([item, item_path])
                

        root_folder = context.web.get_folder_by_server_relative_url(f"{SHAREPOINT_DOC}/{folder_name}")

        for file in file_list:
            print(file)
            file_content = open(file[1], 'rb').read()
            root_folder.upload_file(content=file_content, file_name=file[0]).execute_query()
        
        return (f"File {file[0]} has been uploaded")
    #endregion

    #region Delete Folder
    def delete_folders(self, directories: list):
        for folder_path in directories:
            if os.path.exists(folder_path):
                try:
                    shutil.rmtree(folder_path)
                    print(f"Successfully deleted: {folder_path}")
                except Exception as e:
                    print(f"Error deleting folder '{folder_path}': {e}")
            else:
                print(f"Folder '{folder_path}' does not exist.")
    #endregion

    #region Action 1
    def copy_forecast_from_folder(self, folder_path: str, link_2_folder: str):
        env = environ.Env()
        environ.Env.read_env()

        SHAREPOINT_TO_BE_MACROED = env('SHAREPOINT_TO_BE_MACROED')

        for link_2_filename in os.listdir(link_2_folder):
            if link_2_filename.endswith(".xlsx"):
                link_2_file = os.path.join(link_2_folder, link_2_filename)
                wb_link_2 = openpyxl.load_workbook(link_2_file)

                for filename in os.listdir(folder_path):
                    if filename.endswith(".xlsx"):
                        link_1 = os.path.join(folder_path, filename)

                        wb_link_1 = openpyxl.load_workbook(link_1, data_only=True)
                        sheet_link_1 = wb_link_1.active

                        dealer_name = sheet_link_1.cell(row=4, column=2).value

                        if dealer_name not in wb_link_2.sheetnames:
                            print(f"Dealer sheet '{dealer_name}' not found in {link_2_file}")
                            raise ValueError(f"Dealer sheet '{dealer_name}' not found in {link_2_file}")

                        dealer_sheet_link_2 = wb_link_2[dealer_name]

                        header_row = sheet_link_1[6]
                        forecast_columns = {}

                        for cell in header_row:
                            if str(cell.value).startswith("Forecast"):
                                forecast_columns[cell.value] = cell.column

                        if not forecast_columns:
                            raise ValueError(f"No 'Forecast' columns found in {filename}.")

                        forecast_data = {col_name: [] for col_name in forecast_columns}

                        for row in sheet_link_1.iter_rows(min_row=7):
                            first_cell_value = str(row[0].value)

                            if not first_cell_value.startswith("Total"):
                                for col_name, col_index in forecast_columns.items():
                                    value = row[col_index - 1].value
                                    if value is None or str(value).strip().lower() in ["", "null"]:
                                        value = None
                                    forecast_data[col_name].append(value)

                        for col_name, data in forecast_data.items():
                            col_index_link_2 = None
                            for cell in dealer_sheet_link_2[6]:
                                if cell.value == col_name:
                                    col_index_link_2 = cell.column

                            if col_index_link_2 is None:
                                raise ValueError(f"Column '{col_name}' not found in sheet '{dealer_name}' of {link_2_file}.")

                            for row_idx, value in enumerate(data, start=7):
                                dealer_sheet_link_2.cell(row=row_idx, column=col_index_link_2, value=value)

                        print(f"Forecast data copied for dealer '{dealer_name}' from {filename} to {link_2_file}")

                wb_link_2.save(link_2_file)

        self.upload_excel(link_2_folder, SHAREPOINT_TO_BE_MACROED)
        self.delete_folders([folder_path, link_2_folder])
        return (f"All forecast data copied and uploaded to SharePoint.")
    #endregion
    
    #region Action 2
    def copy_allocations_for_dealers(self, folder_path: str, link_2: str):
        env = environ.Env()
        environ.Env.read_env()

        SHAREPOINT_DEALER_FINAL_ALLOCATION = env('SHAREPOINT_DEALER_FINAL_ALLOCATION')

        dealer_names = [
            "IPN S", "CKMN", "CMC", "DCM", "JIM", "KMS", "KMSA", "MP Solo"
        ]
        
        for filename in os.listdir(folder_path):
            if filename.endswith(".xlsx") and not filename.startswith('~$'):
                link_1 = os.path.join(folder_path, filename)
                wb_obj_1 = openpyxl.load_workbook(link_1, data_only=True)

                print(f"Available sheets in {filename}: {wb_obj_1.sheetnames}")

                for dealer_name in dealer_names:
                    matching_sheets = [sheet for sheet in wb_obj_1.sheetnames if sheet.strip() == dealer_name.strip()]

                    if matching_sheets:
                        
                        allocation_data = self.read_allocation_columns(link_1, matching_sheets[0])

                        print("AAA", allocation_data)
                        
                        matching_file = [f for f in os.listdir(link_2) if dealer_name in f and f.endswith(".xlsx")]
                        
                        if matching_file:
                            link_2_file = os.path.join(link_2, matching_file[0])  # Take the first match
                            self.copy_allocation(dealer_name, allocation_data, link_2_file)
                            print(f"Data copied for dealer '{dealer_name}' to {link_2_file}.")
                        else:
                            print(f"No matching file for dealer '{dealer_name}' in {link_2}.")
                            raise ValueError(f"No matching file for dealer '{dealer_name}' in {link_2}.")
                    else:
                        print(f"Worksheet '{dealer_name}' does not exist in {filename}.")
        
        self.upload_excel(link_2, SHAREPOINT_DEALER_FINAL_ALLOCATION)
        self.delete_folders([folder_path, link_2])
        return ("Processing completed and excel is uploaded to SharePoint")

    def read_allocation_columns(self, link: str, sheet_name: str):
        wb_obj = openpyxl.load_workbook(link, data_only=True)
        sheet_obj = wb_obj[sheet_name]

        header_row = sheet_obj[6]
        forecast_columns = {}

        print("Header row:", [cell.value for cell in header_row])

        for cell in header_row:
            if cell.value and str(cell.value).startswith("Allocation"):
                forecast_columns[cell.value] = cell.column

        print("Identified Forecast Columns:", forecast_columns)

        if not forecast_columns:
            raise ValueError("No columns with headers starting with 'Allocation' found.")

        allocation_data = {col_name: [] for col_name in forecast_columns}
        print("Initialized allocation_data:", allocation_data)

        for row in sheet_obj.iter_rows(min_row=7):
            first_cell_value = row[0].value

            print("First cell value of the row:", first_cell_value)

            if first_cell_value and not str(first_cell_value).startswith("Total"):
                for col_name, col_index in forecast_columns.items():
                    value = row[col_index - 1].value

                    print(f"Value at row {row[0].row}, column {col_index}: {value}")

                    if value is None or isinstance(value, str) and value.strip().lower() in ["", "null"]:
                        value = 0

                    allocation_data[col_name].append(value)

        print("Final allocation_data:", allocation_data)
        return allocation_data

    def copy_allocation(self, sheet_name: str, allocation_data: dict, link_2: str):
        wb_obj_2 = openpyxl.load_workbook(link_2, data_only=False)
        
        sheet_obj_2 = wb_obj_2[sheet_name]
        
        header_row = sheet_obj_2[7]
        
        allocation_columns = [cell.column for cell in header_row if cell.value == "Allocation"]
        
        if not allocation_columns:
            raise ValueError(f"No 'Allocation' columns found in the sheet '{sheet_name}' of the second workbook.")
        
        if len(allocation_columns) != len(allocation_data):
            raise ValueError("Mismatch between the number of 'Allocation' columns in the second Excel and the allocation data.")
        
        for col_idx, (col_name, values) in enumerate(allocation_data.items()):
            target_column = allocation_columns[col_idx]
            
            for row_index, value in enumerate(values, start=8):
                cell_to_update = sheet_obj_2.cell(row=row_index, column=target_column)
                cell_to_update.value = value
        
        wb_obj_2.save(link_2)
        
        return f"Data copied successfully in {link_2}"
    #endregion

    #endregion

    #region Excel Not Used
    def get_dealer_name(self, link_1: str):
        wb_obj = openpyxl.load_workbook(link_1, data_only=True)
        sheet_obj = wb_obj.active
        cell_obj = sheet_obj.cell(row = 4, column = 2)
        return cell_obj.value

    def read_forecast_columns(self, link_1: str):
        wb_obj = openpyxl.load_workbook(link_1, data_only=True)
        sheet_obj = wb_obj.active

        header_row = sheet_obj[6]
        forecast_columns = {}

        for cell in header_row:
            if str(cell.value).startswith("Forecast"):
                forecast_columns[cell.value] = cell.column

        if not forecast_columns:
            raise ValueError("No columns with headers starting with 'Forecast' found.")

        forecast_data = {col_name: [] for col_name in forecast_columns}

        for row in sheet_obj.iter_rows(min_row=7):
            first_cell_value = str(row[0].value)
            
            if not first_cell_value.startswith("Total"):
                for col_name, col_index in forecast_columns.items():
                    value = row[col_index - 1].value

                    if value is None or str(value).strip().lower() in ["", "null"]:
                        value = None

                    forecast_data[col_name].append(value)

        return forecast_data
    
    def copy_forecast(self, sheet_name: str, forecast_data: dict, link_2: str):
        wb_obj_2 = openpyxl.load_workbook(link_2, data_only=False)

        if sheet_name not in wb_obj_2.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the second workbook.")
        
        sheet_obj_2 = wb_obj_2[sheet_name]
        header_row = sheet_obj_2[6]
        
        column_mapping = {cell.value: cell.column for cell in header_row if cell.value is not None}

        for col_name in forecast_data.keys():
            if col_name not in column_mapping:
                found = False
                for cell in header_row:
                    if cell.data_type == 'f' and cell.value == col_name:
                        found = True
                        break
                if not found:
                    raise ValueError(f"Column '{col_name}' not found in the sheet '{sheet_name}' of the second workbook.")

        for col_name, values in forecast_data.items():
            target_column = column_mapping[col_name]

            total_rows = len(values)
            for row_index, value in enumerate(values, start=7):
                if row_index - 7 >= total_rows:
                    break
                cell_to_update = sheet_obj_2.cell(row=row_index, column=target_column)
                cell_to_update.value = value

        wb_obj_2.save(link_2)
        return f"Data copied successfully to '{sheet_name}' in {link_2}"
    
    def get_dealer_name_two(self, link: str):
        wb_obj = openpyxl.load_workbook(link, data_only=True)
        sheet_obj = wb_obj.active
        cell_obj = sheet_obj.cell(row = 3, column = 1)
        dealer_info = cell_obj.value
        if dealer_info and ':' in dealer_info:
            dealer_name = dealer_info.split(':')[1].strip()
            return dealer_name
        return None

    #endregion














